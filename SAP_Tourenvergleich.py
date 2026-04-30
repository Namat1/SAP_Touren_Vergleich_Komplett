import io
from collections import Counter
from typing import Dict, List, Optional, Set, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Stammdaten
# ---------------------------------------------------------------------------

CUSTOMER_GROUPS: Dict[str, List[str]] = {
    "Malchow": [
        "115339", "215634", "216425", "216442", "216467", "216496", "216630", "216133",
        "216432", "216815", "216466", "216615", "219545", "219430", "216590", "215632",
        "216144", "216153", "219208", "216207", "216464", "216529", "216570", "216572",
        "216586", "216588", "216628", "216637", "216744", "219439", "216656", "215551",
        "219544", "216799", "216774", "216122", "216177", "216185", "216221", "216248",
        "216253", "216670", "216672", "219513", "216010", "216178", "216655", "216697",
        "216853", "216653", "216791", "216227", "216290", "216814", "216828", "219427",
        "219570", "216793", "216617", "215014", "215180", "216070", "219586", "216155",
        "216569", "216405", "216623", "219532", "219501", "210650", "216371",
    ],
    "Neumünster": [
        "213406", "214238", "214109", "210353", "211152", "217253", "210750", "210716",
        "214588", "214487", "218394", "210399", "214015", "210492", "218418", "211288",
        "211399", "213095", "218390", "211292", "218373", "218344", "213016", "210234",
        "210276", "218466", "218411", "218420", "218426", "218425", "218468", "218421",
        "214285", "214299", "214297", "214290", "218200", "218711", "218461", "210655",
        "210765", "218355", "210701", "213840", "218208", "211025",
    ],
    "Zarrentin": [
        "213568", "112681", "214289", "213458", "218601", "218804", "214321", "218801",
        "214094", "210509", "213580", "218707", "214376", "211380", "218867", "213553",
        "12823", "214296", "214043", "12923", "214192", "218607", "214590", "210455",
        "214001",
    ],
}

DAY_NAMES = {
    1: "Montag",
    2: "Dienstag",
    3: "Mittwoch",
    4: "Donnerstag",
    5: "Freitag",
    6: "Samstag",
}

# Fallback-Positionen, falls Spaltenüberschriften nicht erkannt werden.
# Index 0 = Spalte A. In der neuen Quelldatei ist:
# A CSB, B SAP, C Name, D Strasse, E Plz, F Ort, G Mo, H Die, I Mitt, J Don, K Fr, L Sam.
DAY_COLUMNS_TOUR = {
    1: 6,
    2: 7,
    3: 8,
    4: 9,
    5: 10,
    6: 11,
}

SAP_COL_INDEX = 0
SAP_DAY_COL_INDEX = 6
TOUR_SAP_COL_INDEX = 1

# Diese Blätter werden aus der Quelldatei gelesen.
# Alles andere wie LADEREIHENFOLGE, KUNDENDATEN, TBS_DRUCK usw. wird ignoriert.
TOUR_SHEET_CANDIDATES = [
    "DIREKT",
    "MK",
    "HUPA_NMS",
    "HUPA_MALCHOW",
]

# Neue kurze Tagesüberschriften aus der Quelldatei.
DAY_COLUMN_CANDIDATES = {
    1: ["mo", "montag"],
    2: ["die", "di", "dienstag"],
    3: ["mitt", "mit", "mi", "mittwoch"],
    4: ["don", "do", "donnerstag"],
    5: ["fr", "frei", "freitag"],
    6: ["sam", "sa", "samstag"],
}

LOCATION_ORDER = {name: index for index, name in enumerate(CUSTOMER_GROUPS.keys(), start=1)}
CUSTOMER_TO_LOCATION: Dict[str, str] = {}
CUSTOMER_TO_ORDER: Dict[str, int] = {}

for location_name, sap_list in CUSTOMER_GROUPS.items():
    for customer_index, sap_number in enumerate(sap_list, start=1):
        CUSTOMER_TO_LOCATION[sap_number] = location_name
        CUSTOMER_TO_ORDER[sap_number] = customer_index

SELECTED_SAPS: Set[str] = set(CUSTOMER_TO_LOCATION.keys())


# ---------------------------------------------------------------------------
# Helfer
# ---------------------------------------------------------------------------

def find_duplicate_saps() -> List[Tuple[str, List[str]]]:
    """Findet SAP-Nummern, die in mehreren Standorten oder mehrfach im selben
    Standort hinterlegt sind."""
    counter: Counter = Counter()
    location_map: Dict[str, List[str]] = {}
    for location, sap_list in CUSTOMER_GROUPS.items():
        for sap in sap_list:
            counter[sap] += 1
            location_map.setdefault(sap, []).append(location)
    return [(sap, location_map[sap]) for sap, count in counter.items() if count > 1]


def normalize_sap_series(series: pd.Series) -> pd.Series:
    """Normalisiert SAP-Nummern vektorisiert. Floats ohne Nachkomma werden zu int."""
    if series.empty:
        return series.astype(str)

    result = series.copy()
    # Numerische Werte -> int wenn möglich
    numeric = pd.to_numeric(result, errors="coerce")
    is_int = numeric.notna() & (numeric == numeric.round())

    out = result.astype(str)
    out = out.where(~is_int, numeric.where(is_int).astype("Int64").astype(str))
    out = out.str.strip()
    out = out.replace({"nan": "", "<NA>": "", "None": ""})
    return out


def normalize_header_name(value) -> str:
    """Vereinfacht Spaltenüberschriften für robuste Erkennung."""
    text = "" if value is None or pd.isna(value) else str(value)
    text = text.strip().lower()
    text = (
        text.replace("ä", "ae")
        .replace("ö", "oe")
        .replace("ü", "ue")
        .replace("ß", "ss")
    )
    return "".join(ch for ch in text if ch.isalnum())


def pick_first_matching_column(columns: List[str], candidates: List[str]) -> Optional[str]:
    candidate_set = set(candidates)
    for column in columns:
        if normalize_header_name(column) in candidate_set:
            return column
    return None


def make_unique_columns(raw_columns: List[object]) -> List[str]:
    """Erzeugt eindeutige Spaltennamen, auch wenn Excel leere/gleiche Überschriften enthält."""
    result: List[str] = []
    seen: Dict[str, int] = {}
    for index, value in enumerate(raw_columns, start=1):
        name = value_to_clean_text(value)
        if not name:
            name = f"Spalte_{index}"
        count = seen.get(name, 0) + 1
        seen[name] = count
        if count > 1:
            name = f"{name}_{count}"
        result.append(name)
    return result


def normalized_candidates(values: List[str]) -> List[str]:
    return [normalize_header_name(value) for value in values]


def pick_column_by_name_or_position(
    columns: List[str],
    candidates: List[str],
    fallback_index: Optional[int] = None,
) -> Optional[str]:
    """Findet eine Spalte zuerst per Überschrift, sonst per alter Position."""
    found = pick_first_matching_column(columns, normalized_candidates(candidates))
    if found is not None:
        return found
    if fallback_index is not None and len(columns) > fallback_index:
        return columns[fallback_index]
    return None


def read_excel_with_detected_header(excel: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """Liest ein Blatt und erkennt die Kopfzeile selbst.

    Das ist wichtig, wenn oberhalb der eigentlichen Tabelle noch Leerzeilen,
    Filterzeilen oder Druck-Kopfbereiche stehen. Gesucht wird eine Zeile mit
    SAP und mindestens zwei Tages-Spalten wie Mo, Die, Mitt, Don, Fr, Sam.
    """
    raw = pd.read_excel(excel, sheet_name=sheet_name, header=None, dtype=object)
    if raw.empty:
        return pd.DataFrame()

    header_row: Optional[int] = None
    max_scan_rows = min(len(raw), 25)
    day_names_flat = {
        normalize_header_name(candidate)
        for values in DAY_COLUMN_CANDIDATES.values()
        for candidate in values
    }

    for row_index in range(max_scan_rows):
        values = [normalize_header_name(value) for value in raw.iloc[row_index].tolist()]
        value_set = set(values)
        has_sap = "sap" in value_set or "sapnummer" in value_set or "sapnr" in value_set
        day_hits = sum(1 for value in values if value in day_names_flat)
        if has_sap and day_hits >= 2:
            header_row = row_index
            break

    if header_row is None:
        # Fallback: wie vorher mit erster Zeile als Kopfzeile.
        return pd.read_excel(excel, sheet_name=sheet_name, header=0, dtype=object)

    df = raw.iloc[header_row + 1:].copy()
    df.columns = make_unique_columns(raw.iloc[header_row].tolist())
    df = df.dropna(how="all").reset_index(drop=True)
    return df


def select_tour_sheet_names(excel: pd.ExcelFile) -> List[str]:
    """Wählt nur die echten Quelldatei-Blätter aus."""
    available_by_normalized = {normalize_header_name(name): name for name in excel.sheet_names}
    selected: List[str] = []

    for expected in TOUR_SHEET_CANDIDATES:
        real_name = available_by_normalized.get(normalize_header_name(expected))
        if real_name and real_name not in selected:
            selected.append(real_name)

    if selected:
        return selected

    # Fallback, falls die Datei anders benannt wurde: die ersten vier Blätter wie bisher.
    return excel.sheet_names[:4]


def day_value_is_set(value) -> bool:
    """Ein Tag gilt als vorhanden, wenn in Mo/Die/Mitt/Don/Fr/Sam irgendein Wert steht.

    Tournummern wie 1028, 2063, 3028 zählen. 0, leere Zellen und Striche zählen nicht.
    """
    if value is None or pd.isna(value):
        return False
    text = str(value).strip()
    if not text:
        return False
    if text.lower() in {"nan", "none", "<na>", "-", "--"}:
        return False
    number = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    if pd.notna(number) and float(number) == 0:
        return False
    return True


def normalize_day_code_series(series: pd.Series) -> pd.Series:
    """Normalisiert Liefertage aus SAP: 1..6 oder Text wie Montag/Mo."""
    numeric = pd.to_numeric(series, errors="coerce")
    text = series.astype(str).map(normalize_header_name)
    text_map = {
        "1": 1, "mo": 1, "montag": 1,
        "2": 2, "di": 2, "die": 2, "dienstag": 2,
        "3": 3, "mi": 3, "mitt": 3, "mit": 3, "mittwoch": 3,
        "4": 4, "do": 4, "don": 4, "donnerstag": 4,
        "5": 5, "fr": 5, "frei": 5, "freitag": 5,
        "6": 6, "sa": 6, "sam": 6, "samstag": 6,
    }
    mapped = text.map(text_map)
    return numeric.where(numeric.notna(), mapped)


def value_to_clean_text(value) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def merge_customer_info(base: Dict[str, Dict[str, str]], sap: str, info: Dict[str, str]) -> None:
    target = base.setdefault(sap, {"name": "", "strasse": "", "ort": ""})
    for key in ["name", "strasse", "ort"]:
        if not target.get(key) and info.get(key):
            target[key] = info[key]


def read_sap_file(uploaded_file) -> Tuple[Dict[str, Set[int]], str, int]:
    """Liest die SAP-Datei.

    Bevorzugt werden Spaltenüberschriften wie SAP und Liefertag. Falls die
    Überschriften anders heißen, wird auf die alte Struktur zurückgefallen:
    Spalte A = SAP, Spalte G = Liefertag 1 bis 6.
    """
    excel = pd.ExcelFile(uploaded_file)
    sheet_name = excel.sheet_names[0]
    df = read_excel_with_detected_header(excel, sheet_name)

    if df.empty:
        return {}, sheet_name, 0

    columns = list(df.columns)
    sap_column = pick_column_by_name_or_position(
        columns,
        ["SAP", "SAP Nummer", "SAP-Nr", "SAP Nr", "Kundennummer", "Kunden Nummer"],
        SAP_COL_INDEX,
    )
    day_column = pick_column_by_name_or_position(
        columns,
        ["Liefertag", "Liefer Tag", "LT", "Tag", "Liefertag Code", "Liefertagcode"],
        SAP_DAY_COL_INDEX,
    )

    if sap_column is None or day_column is None:
        return {}, sheet_name, 0

    work = df[[sap_column, day_column]].copy()
    work.columns = ["sap", "tag"]
    work["sap"] = normalize_sap_series(work["sap"])
    work["tag_num"] = normalize_day_code_series(work["tag"])

    mask = (
        work["sap"].ne("")
        & work["tag_num"].between(1, 6, inclusive="both")
        & work["tag_num"].notna()
    )
    filtered = work.loc[mask, ["sap", "tag_num"]].copy()
    filtered["tag_int"] = filtered["tag_num"].astype(int)

    days_by_sap: Dict[str, Set[int]] = (
        filtered.groupby("sap")["tag_int"].agg(set).to_dict()
    )

    return days_by_sap, sheet_name, len(filtered)


def read_tourenplanung(uploaded_file) -> Tuple[pd.DataFrame, List[str], Dict[str, Dict[str, str]]]:
    """Liest die neue Quelldatei.

    Erwartet werden die Blätter DIREKT, MK, HUPA_NMS und HUPA_MALCHOW.
    Die Spalten werden per Name erkannt:
    CSB, SAP, Name, Strasse, Plz, Ort, Mo, Die, Mitt, Don, Fr, Sam.

    Zusätzlich bleiben die alten Positionen als Fallback erhalten:
    B = SAP und G bis L = Montag bis Samstag.
    """
    excel = pd.ExcelFile(uploaded_file)
    sheet_names = select_tour_sheet_names(excel)

    frames: List[pd.DataFrame] = []
    customer_info: Dict[str, Dict[str, str]] = {}

    for sheet_name in sheet_names:
        df = read_excel_with_detected_header(excel, sheet_name)
        if df.empty:
            continue

        columns = list(df.columns)
        sap_column = pick_column_by_name_or_position(
            columns,
            ["SAP", "SAP Nummer", "SAP-Nr", "SAP Nr", "Kundennummer", "Kunden Nummer"],
            TOUR_SAP_COL_INDEX,
        )
        if sap_column is None:
            continue

        csb_column = pick_column_by_name_or_position(columns, ["CSB", "CSB Nummer", "CSB-Nr", "CSB Nr"], 0)
        name_column = pick_column_by_name_or_position(
            columns,
            ["Name", "Kundenname", "Marktname", "Kunde", "Bezeichnung", "Filialname"],
            2,
        )
        strasse_column = pick_column_by_name_or_position(
            columns,
            ["Strasse", "Straße", "Str", "Anschrift", "Adresse", "Strassenname", "Straßenname", "Strasse Hausnummer", "Straße Hausnummer"],
            3,
        )
        plz_column = pick_column_by_name_or_position(columns, ["Plz", "PLZ", "Postleitzahl"], 4)
        ort_column = pick_column_by_name_or_position(columns, ["Ort", "Stadt", "Plz Ort", "PLZ Ort", "Ortname"], 5)

        rename_map = {sap_column: "sap"}
        if csb_column and csb_column != sap_column:
            rename_map[csb_column] = "csb"

        for day_num, col_index in DAY_COLUMNS_TOUR.items():
            day_column = pick_column_by_name_or_position(
                columns,
                DAY_COLUMN_CANDIDATES[day_num],
                col_index,
            )
            if day_column and day_column != sap_column:
                rename_map[day_column] = f"tag_{day_num}"

        if name_column and name_column != sap_column:
            rename_map[name_column] = "name"
        if strasse_column and strasse_column != sap_column:
            rename_map[strasse_column] = "strasse"
        if ort_column and ort_column != sap_column:
            rename_map[ort_column] = "ort"
        if plz_column and plz_column != sap_column:
            rename_map[plz_column] = "plz"

        work = df.rename(columns=rename_map).copy()
        work["sap"] = normalize_sap_series(work["sap"])
        work = work[work["sap"].ne("")].copy()
        if work.empty:
            continue

        if "ort" not in work.columns:
            work["ort"] = ""
        if "plz" not in work.columns:
            work["plz"] = ""
        if "name" not in work.columns:
            work["name"] = ""
        if "strasse" not in work.columns:
            work["strasse"] = ""

        info_cols = ["sap", "name", "strasse", "ort", "plz"]
        info_df = work[info_cols].copy()
        info_df["name"] = info_df["name"].map(value_to_clean_text)
        info_df["strasse"] = info_df["strasse"].map(value_to_clean_text)
        info_df["ort"] = info_df["ort"].map(value_to_clean_text)
        info_df["plz"] = info_df["plz"].map(value_to_clean_text)
        info_df["ort_kombi"] = info_df.apply(
            lambda row: " ".join(v for v in [row["plz"], row["ort"]] if v).strip(),
            axis=1,
        )

        for _, row in info_df.iterrows():
            merge_customer_info(
                customer_info,
                row["sap"],
                {
                    "name": row["name"],
                    "strasse": row["strasse"],
                    "ort": row["ort_kombi"] or row["ort"],
                },
            )

        day_value_columns = [f"tag_{d}" for d in DAY_COLUMNS_TOUR.keys() if f"tag_{d}" in work.columns]
        if not day_value_columns:
            continue

        work["blatt"] = sheet_name
        long = work.melt(
            id_vars=["sap", "blatt"],
            value_vars=day_value_columns,
            var_name="tag_col",
            value_name="wert",
        )
        long["tag_num"] = long["tag_col"].str.replace("tag_", "", regex=False).astype(int)
        long["wert_gesetzt"] = long["wert"].map(day_value_is_set)

        long = long[long["sap"].ne("") & long["wert_gesetzt"]]
        frames.append(long[["sap", "blatt", "tag_num", "wert"]])

    if not frames:
        return pd.DataFrame(columns=["sap", "blatt", "tag_num", "wert"]), sheet_names, customer_info

    return pd.concat(frames, ignore_index=True), sheet_names, customer_info


def build_missing_in_sap(
    tour_df: pd.DataFrame,
    days_by_sap: Dict[str, Set[int]],
    customer_info: Dict[str, Dict[str, str]],
) -> pd.DataFrame:
    """Eine Zeile pro Kunde: welche Tage stehen in der Tourenplanung, fehlen aber
    in SAP als Liefertag. Keine Restriktion auf Kundengruppen – alle SAP-Nummern."""
    if tour_df.empty:
        return _empty_result_df()

    known = tour_df[tour_df["sap"].ne("")].copy()
    if known.empty:
        return _empty_result_df()

    known["fehlt"] = known.apply(
        lambda row: row["tag_num"] not in days_by_sap.get(row["sap"], set()),
        axis=1,
    )
    missing = known[known["fehlt"]]
    if missing.empty:
        return _empty_result_df()

    days_in_tour: Dict[str, Set[int]] = (
        tour_df.groupby("sap")["tag_num"].agg(set).to_dict()
    )

    agg = missing.groupby("sap", as_index=False).agg(
        tage=("tag_num", lambda x: sorted(set(x))),
    )

    agg["Standort"] = agg["sap"].map(CUSTOMER_TO_LOCATION).fillna("Direkt")
    agg["Name"] = agg["sap"].map(lambda s: customer_info.get(s, {}).get("name", ""))
    agg["Straße"] = agg["sap"].map(lambda s: customer_info.get(s, {}).get("strasse", ""))
    agg["Ort"] = agg["sap"].map(lambda s: customer_info.get(s, {}).get("ort", ""))
    agg["Fehlende LT"] = agg["tage"].map(
        lambda tage: ", ".join(f"{d} {DAY_NAMES[d]}" for d in tage)
    )
    agg["LT SAP"] = agg["sap"].map(
        lambda s: ", ".join(f"{d} {DAY_NAMES[d]}" for d in sorted(days_by_sap.get(s, set())))
        or "(keine hinterlegt)"
    )
    agg["LT Tourenplanung"] = agg["sap"].map(
        lambda s: ", ".join(f"{d} {DAY_NAMES[d]}" for d in sorted(days_in_tour.get(s, set())))
    )

    agg["_HupaFlag"] = agg["sap"].isin(SELECTED_SAPS)
    agg["_StandortSort"] = agg["Standort"].map(LOCATION_ORDER).fillna(999)
    agg["_SapSort"] = pd.to_numeric(agg["sap"], errors="coerce").fillna(9_999_999_999)
    agg = agg.rename(columns={"sap": "SAP Nummer"}).sort_values(
        ["_StandortSort", "_SapSort"]
    ).reset_index(drop=True)

    keep = _export_columns_missing() + ["_HupaFlag"]
    return agg[keep]


def build_missing_in_tour(
    tour_df: pd.DataFrame,
    days_by_sap: Dict[str, Set[int]],
    customer_info: Dict[str, Dict[str, str]],
) -> pd.DataFrame:
    """Eine Zeile pro Kunde: welche Tage sind in SAP als Liefertag hinterlegt,
    fehlen aber in der Tourenplanung. Keine Restriktion auf Kundengruppen."""
    days_in_tour: Dict[str, Set[int]] = {}
    if not tour_df.empty:
        days_in_tour = tour_df.groupby("sap")["tag_num"].agg(set).to_dict()

    rows: List[dict] = []
    for sap, sap_days in days_by_sap.items():
        tour_days = days_in_tour.get(sap, set())
        fehlend = sorted(sap_days - tour_days)
        if not fehlend:
            continue
        standort = CUSTOMER_TO_LOCATION.get(sap, "Direkt")
        info = customer_info.get(sap, {})
        rows.append({
            "Standort": standort,
            "SAP Nummer": sap,
            "Name": info.get("name", ""),
            "Straße": info.get("strasse", ""),
            "Ort": info.get("ort", ""),
            "Fehlende LT": ", ".join(f"{d} {DAY_NAMES[d]}" for d in fehlend),
            "LT SAP": ", ".join(
                f"{d} {DAY_NAMES[d]}" for d in sorted(sap_days)
            ),
            "LT Tourenplanung": ", ".join(
                f"{d} {DAY_NAMES[d]}" for d in sorted(tour_days)
            ) or "(nicht in Tourenplanung vorhanden)",
            "_HupaFlag": sap in SELECTED_SAPS,
            "_StandortSort": LOCATION_ORDER.get(standort, 999),
            "_SapSort": int(sap) if sap.isdigit() else 9_999_999_999,
        })

    if not rows:
        return pd.DataFrame(columns=_export_columns_missing_tour() + ["_HupaFlag"])

    df = pd.DataFrame(rows)
    df = df.sort_values(["_StandortSort", "_SapSort"]).reset_index(drop=True)
    return df[_export_columns_missing_tour() + ["_HupaFlag"]]


def split_hupa_direkt(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Teilt ein Ergebnis-DataFrame in Hupa (SAPs aus CUSTOMER_GROUPS) und
    Direkt (alles andere). Für Direkt wird die Standort-Spalte entfernt und
    nach SAP-Nummer sortiert."""
    if df.empty:
        return df.drop(columns=["_HupaFlag"], errors="ignore"), df.drop(columns=["_HupaFlag", "Standort"], errors="ignore")

    hupa = df[df["_HupaFlag"]].drop(columns=["_HupaFlag"]).reset_index(drop=True)
    direkt = df[~df["_HupaFlag"]].drop(columns=["_HupaFlag", "Standort"], errors="ignore")
    # Direkt nach SAP-Nummer numerisch sortieren
    direkt = direkt.assign(
        _s=pd.to_numeric(direkt["SAP Nummer"], errors="coerce").fillna(9_999_999_999)
    ).sort_values("_s").drop(columns=["_s"]).reset_index(drop=True)
    return hupa, direkt


def _export_columns_missing() -> List[str]:
    return [
        "Standort",
        "SAP Nummer",
        "Name",
        "Straße",
        "Ort",
        "Fehlende LT",
        "LT SAP",
        "LT Tourenplanung",
    ]


def _export_columns_missing_tour() -> List[str]:
    return [
        "Standort",
        "SAP Nummer",
        "Name",
        "Straße",
        "Ort",
        "Fehlende LT",
        "LT SAP",
        "LT Tourenplanung",
    ]


def _empty_result_df() -> pd.DataFrame:
    return pd.DataFrame(columns=_export_columns_missing() + ["_HupaFlag"])


def _add_count_column(df: pd.DataFrame) -> pd.DataFrame:
    """Fügt eine Hilfsspalte 'Anzahl LT' (Anzahl fehlender Liefertage) hinzu
    und positioniert sie hinter 'SAP Nummer'."""
    if df is None or df.empty:
        return df
    out = df.copy()
    out["Anzahl LT"] = out["Fehlende LT"].fillna("").map(
        lambda s: len([t for t in str(s).split(",") if t.strip()])
    )
    cols = list(out.columns)
    cols.remove("Anzahl LT")
    if "SAP Nummer" in cols:
        idx = cols.index("SAP Nummer") + 1
        cols.insert(idx, "Anzahl LT")
    else:
        cols.insert(0, "Anzahl LT")
    return out[cols]


def _filter_dataframe(df: pd.DataFrame, suche: str, standort: Optional[str] = None) -> pd.DataFrame:
    """Filtert nach freitext (SAP, Name, Straße, Ort) und optional Standort."""
    if df is None or df.empty:
        return df
    work = df
    if standort and standort != "Alle" and "Standort" in work.columns:
        work = work[work["Standort"] == standort]
    if suche:
        such = suche.strip().lower()
        if such:
            spalten = [c for c in ["SAP Nummer", "Name", "Straße", "Ort"] if c in work.columns]
            mask = pd.Series(False, index=work.index)
            for c in spalten:
                mask = mask | work[c].astype(str).str.lower().str.contains(such, na=False)
            work = work[mask]
    return work


def _standort_uebersicht(hupa_sap: pd.DataFrame, hupa_tour: Optional[pd.DataFrame]) -> pd.DataFrame:
    """Aggregiert pro Standort: Anzahl betroffener Kunden in beiden Richtungen."""
    rows = []
    standorte = list(CUSTOMER_GROUPS.keys())
    for standort in standorte:
        gesamt = len(CUSTOMER_GROUPS[standort])
        betr_sap = (
            int((hupa_sap["Standort"] == standort).sum())
            if hupa_sap is not None and not hupa_sap.empty and "Standort" in hupa_sap.columns
            else 0
        )
        betr_tour = (
            int((hupa_tour["Standort"] == standort).sum())
            if hupa_tour is not None and not hupa_tour.empty and "Standort" in hupa_tour.columns
            else 0
        )
        rows.append({
            "Standort": standort,
            "Kunden gesamt": gesamt,
            "Fehlt in SAP": betr_sap,
            "Fehlt in Tour": betr_tour,
        })
    return pd.DataFrame(rows)


def build_excel(
    missing_sap: pd.DataFrame,
    missing_tour: pd.DataFrame | None,
) -> bytes:
    """Schreibt eine Excel mit je einem Blatt 'Hupa' und 'Direkt' für die
    Standard-Richtung (Fehlt in SAP) und optional den zwei Reverse-Blättern."""
    hupa_sap, direkt_sap = split_hupa_direkt(missing_sap)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        hupa_out = _add_count_column(hupa_sap)
        hupa_out.to_excel(writer, index=False, sheet_name="Hupa", na_rep="")
        _format_sheet(writer, "Hupa", hupa_out)

        direkt_out = _add_count_column(direkt_sap)
        direkt_out.to_excel(writer, index=False, sheet_name="Direkt", na_rep="")
        _format_sheet(writer, "Direkt", direkt_out)

        if missing_tour is not None:
            hupa_tour, direkt_tour = split_hupa_direkt(missing_tour)

            hupa_tour_out = _add_count_column(hupa_tour)
            hupa_tour_out.to_excel(writer, index=False, sheet_name="Hupa - Fehlt in Tour", na_rep="")
            _format_sheet(writer, "Hupa - Fehlt in Tour", hupa_tour_out)

            direkt_tour_out = _add_count_column(direkt_tour)
            direkt_tour_out.to_excel(writer, index=False, sheet_name="Direkt - Fehlt in Tour", na_rep="")
            _format_sheet(writer, "Direkt - Fehlt in Tour", direkt_tour_out)

        # Defensiv: alle Sheets sichtbar; falls openpyxl ein Default-Sheet "Sheet"
        # angelegt hat und es noch leer ist, entfernen.
        wb = writer.book
        for ws in list(wb.worksheets):
            if ws.title == "Sheet" and ws.max_row == 1 and ws.max_column == 1:
                wb.remove(ws)
        for ws in wb.worksheets:
            ws.sheet_state = "visible"
        if not wb.worksheets:
            wb.create_sheet("Hupa")

    return output.getvalue()


# Spalten, die rechtsbündig dargestellt werden (Zahlen)
_RIGHT_ALIGN_COLS = {"SAP Nummer", "Anzahl LT"}

# Empfohlene Mindest-/Maximal-Breiten je Spalte
_COL_WIDTH_HINTS = {
    "Standort": (12, 16),
    "SAP Nummer": (10, 12),
    "Anzahl LT": (10, 11),
    "Name": (24, 42),
    "Straße": (20, 32),
    "Ort": (22, 36),
    "Fehlende LT": (24, 48),
    "LT SAP": (24, 48),
    "LT Tourenplanung": (24, 48),
}


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame) -> None:
    if df is None:
        return
    ws = writer.sheets[sheet_name]
    n_rows = len(df)
    n_cols = len(df.columns)

    # Farben
    header_fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")
    zebra_fill = PatternFill(start_color="FFE8EFF7", end_color="FFE8EFF7", fill_type="solid")

    # Borders
    thin = Side(style="thin", color="FFCBD5E0")
    medium = Side(style="medium", color="FF305496")
    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
    body_font = Font(name="Calibri", size=11)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=False)
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Header formatieren
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_left
        cell.border = border_thin
    ws.row_dimensions[1].height = 24

    # Standort-Spalte ermitteln (für dicke Trennlinie bei Wechsel)
    columns = list(df.columns)
    standort_idx = columns.index("Standort") if "Standort" in columns else None
    standort_values = df["Standort"].tolist() if standort_idx is not None else []

    # Datenzeilen formatieren
    for row_offset in range(n_rows):
        excel_row = row_offset + 2
        is_zebra = (row_offset % 2) == 1

        # Trennt sich der Standort zur Vorzeile? -> dicke obere Linie
        new_group = False
        if standort_idx is not None and row_offset > 0:
            if standort_values[row_offset] != standort_values[row_offset - 1]:
                new_group = True

        ws.row_dimensions[excel_row].height = 20

        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.font = body_font

            if col_name == "Anzahl LT":
                cell.alignment = align_center
            elif col_name in _RIGHT_ALIGN_COLS:
                cell.alignment = align_right
            else:
                cell.alignment = align_left

            if is_zebra:
                cell.fill = zebra_fill

            top_side = medium if new_group else thin
            cell.border = Border(left=thin, right=thin, top=top_side, bottom=thin)

    # Spaltenbreiten: Hinweise + Inhaltslänge
    for col_idx, col_name in enumerate(columns, start=1):
        sample = df[col_name].astype(str).head(300).tolist()
        max_len = max([len(str(col_name))] + [len(v) for v in sample] + [8])
        min_w, max_w = _COL_WIDTH_HINTS.get(col_name, (12, 50))
        width = min(max(max_len + 3, min_w), max_w)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Kopfzeile einfrieren + Autofilter
    ws.freeze_panes = "A2"
    if n_rows > 0:
        last_col = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"

    # Conditional Formatting für 'Anzahl LT': je höher, desto roter
    if "Anzahl LT" in columns and n_rows > 0:
        from openpyxl.formatting.rule import ColorScaleRule
        col_letter = get_column_letter(columns.index("Anzahl LT") + 1)
        rng = f"{col_letter}2:{col_letter}{n_rows + 1}"
        rule = ColorScaleRule(
            start_type="num", start_value=1, start_color="FFD4EDDA",
            mid_type="num", mid_value=3, mid_color="FFFFE699",
            end_type="num", end_value=6, end_color="FFF4B084",
        )
        ws.conditional_formatting.add(rng, rule)

    # Druck: Querformat und Kopfzeile auf jeder Seite. Defensiv ohne Properties,
    # die in manchen openpyxl-Versionen "At least one sheet must be visible" auslösen.
    try:
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.print_options.gridLines = False
        if n_rows > 0:
            ws.print_title_rows = "1:1"
    except Exception:
        # Druckeinstellungen sind nice-to-have, blockieren aber niemals den Export
        pass

    # Sicherstellen, dass das Sheet sichtbar ist (sonst kann openpyxl
    # beim Speichern "At least one sheet must be visible" werfen)
    ws.sheet_state = "visible"


def build_group_overview() -> str:
    parts: List[str] = []
    for location_name, sap_list in CUSTOMER_GROUPS.items():
        parts.append(f"{location_name} ({len(sap_list)} Kunden)")
        parts.append("\n".join(sap_list))
        parts.append("")
    return "\n".join(parts).strip()


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Tourenplanung gegen SAP (Hupa + Direkt)", layout="wide")

st.title("Tourenplanung gegen SAP – alle SAP-Nummern")
st.write(
    "Vergleicht die Liefertage in der Tourenplanung gegen die in SAP hinterlegten Liefertage "
    "für alle SAP-Nummern ohne Filter. "
    "Name, Straße und Ort werden dabei aus der Tourenplanung gelesen. "
    "Kunden, die in den hinterlegten Standorten (Malchow, Neumünster, Zarrentin) gelistet sind, "
    "landen im Blatt **Hupa**. Alle anderen im Blatt **Direkt**."
)

# Datenqualitäts-Warnung bei Duplikaten in der Konfiguration
duplicates = find_duplicate_saps()
if duplicates:
    with st.expander(f"⚠️ {len(duplicates)} doppelte SAP-Nummer(n) in der Kundensortierung", expanded=False):
        for sap, locations in duplicates:
            st.write(f"- **{sap}**: {', '.join(locations)}")

st.info(
    "Richtung des Vergleichs:\n"
    "- SAP = Datei mit SAP Nummer und Liefertag 1 bis 6. Falls die Überschriften anders sind, wird A = SAP und G = Liefertag genutzt.\n"
    "- Quelldatei = Blätter DIREKT, MK, HUPA_NMS und HUPA_MALCHOW.\n"
    "- Quelldatei-Spalten = CSB, SAP, Name, Strasse, Plz, Ort, Mo, Die, Mitt, Don, Fr, Sam.\n"
    "- Ein Wert in Mo bis Sam zählt als Lieferung an diesem Tag.\n"
    "- Ausgabe = zwei Blätter, Hupa (bekannte Standorte) und Direkt (Rest)."
)

col1, col2, col3 = st.columns(3)
col1.metric("Malchow", len(CUSTOMER_GROUPS["Malchow"]))
col2.metric("Neumünster", len(CUSTOMER_GROUPS["Neumünster"]))
col3.metric("Zarrentin", len(CUSTOMER_GROUPS["Zarrentin"]))

with st.expander("Hinterlegte Kundenliste (Hupa)", expanded=False):
    st.text_area(
        "Diese SAP-Nummern landen im Blatt Hupa. Alles andere im Blatt Direkt.",
        value=build_group_overview(),
        height=420,
        disabled=True,
    )

sap_datei = st.file_uploader(
    "SAP hochladen – erstes Blatt, Spalte A = SAP Nummer, G = Liefertag 1 bis 6",
    type=["xlsx", "xlsm", "xls"],
    key="sap_datei",
)

tourenplanung_datei = st.file_uploader(
    "Quelldatei hochladen – Blätter DIREKT, MK, HUPA_NMS, HUPA_MALCHOW; Spalten CSB, SAP, Name, Strasse, Plz, Ort, Mo, Die, Mitt, Don, Fr, Sam",
    type=["xlsx", "xlsm", "xls"],
    key="tourenplanung_datei",
)

with st.expander("Optionen", expanded=False):
    include_reverse = st.checkbox(
        "Zusätzlich prüfen: Tage, die in SAP stehen, aber in der Tourenplanung fehlen "
        "(zwei weitere Blätter: 'Hupa - Fehlt in Tour', 'Direkt - Fehlt in Tour')",
        value=False,
    )

run = st.button("Excel erzeugen", type="primary")

if run:
    if not sap_datei or not tourenplanung_datei:
        st.error("Bitte beide Excel-Dateien hochladen.")
        st.stop()

    try:
        days_by_sap, sap_sheet, sap_rows = read_sap_file(sap_datei)
        tour_df, tour_sheets, customer_info = read_tourenplanung(tourenplanung_datei)

        if sap_rows == 0:
            st.warning("In der SAP-Datei wurden keine gültigen Liefertage erkannt. Erwartet wird SAP Nummer und Liefertag 1 bis 6.")
        if tour_df.empty:
            st.warning(
                "In der Quelldatei wurden keine gesetzten Liefertage erkannt. "
                f"Geprüfte Blätter: {', '.join(tour_sheets)}. "
                "Erwartete Spalten: CSB, SAP, Name, Strasse, Plz, Ort, Mo, Die, Mitt, Don, Fr, Sam."
            )

        missing_sap = build_missing_in_sap(tour_df, days_by_sap, customer_info)
        missing_tour = build_missing_in_tour(tour_df, days_by_sap, customer_info) if include_reverse else None

        excel_bytes = build_excel(missing_sap, missing_tour)

        hupa_sap, direkt_sap = split_hupa_direkt(missing_sap)
        hupa_tour, direkt_tour = (split_hupa_direkt(missing_tour) if missing_tour is not None else (None, None))

        st.session_state["result"] = {
            "hupa_sap": hupa_sap,
            "direkt_sap": direkt_sap,
            "hupa_tour": hupa_tour,
            "direkt_tour": direkt_tour,
            "excel_bytes": excel_bytes,
            "sap_sheet": sap_sheet,
            "sap_rows": sap_rows,
            "tour_sheets": tour_sheets,
            "tour_rows": len(tour_df),
        }
    except Exception as exc:
        import traceback
        st.error(f"Fehler beim Verarbeiten der Dateien: {exc}")
        with st.expander("Technische Details", expanded=False):
            st.code(traceback.format_exc(), language="python")
        st.session_state.pop("result", None)

# ---------------------------------------------------------------------------
# Ergebnisanzeige
# ---------------------------------------------------------------------------

result = st.session_state.get("result")
if result:
    hupa_sap = result["hupa_sap"]
    direkt_sap = result["direkt_sap"]
    hupa_tour = result["hupa_tour"]
    direkt_tour = result["direkt_tour"]
    has_reverse = hupa_tour is not None or direkt_tour is not None

    st.divider()

    # Kopfzeile: Kennzahlen + Download
    head_left, head_right = st.columns([3, 1])
    with head_left:
        st.subheader("Ergebnis")
        st.caption(
            f"SAP: Blatt **{result['sap_sheet']}**, {result['sap_rows']} Liefertage übernommen · "
            f"Quelldatei: {', '.join(result['tour_sheets'])}, {result.get('tour_rows', 0)} gesetzte Liefertage erkannt"
        )
    with head_right:
        st.download_button(
            label="📥 Excel herunterladen",
            data=result["excel_bytes"],
            file_name="tourenplanung_sap_hupa_direkt.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    # Hauptkennzahlen
    if has_reverse:
        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Hupa – Fehlt in SAP", len(hupa_sap) if hupa_sap is not None else 0)
        m2.metric("Direkt – Fehlt in SAP", len(direkt_sap) if direkt_sap is not None else 0)
        m3.metric("Hupa – Fehlt in Tour", len(hupa_tour) if hupa_tour is not None else 0)
        m4.metric("Direkt – Fehlt in Tour", len(direkt_tour) if direkt_tour is not None else 0)
    else:
        m1, m2, m3 = st.columns(3)
        m1.metric("Hupa betroffen", len(hupa_sap) if hupa_sap is not None else 0)
        m2.metric("Direkt betroffen", len(direkt_sap) if direkt_sap is not None else 0)
        m3.metric("Summe", (len(hupa_sap) if hupa_sap is not None else 0) + (len(direkt_sap) if direkt_sap is not None else 0))

    # Tabs für die Detailansichten
    tab_labels = [
        "📊 Übersicht",
        f"🟢 Hupa · Fehlt in SAP ({len(hupa_sap)})",
        f"🟡 Direkt · Fehlt in SAP ({len(direkt_sap)})",
    ]
    if has_reverse:
        tab_labels.append(f"🔵 Hupa · Fehlt in Tour ({len(hupa_tour) if hupa_tour is not None else 0})")
        tab_labels.append(f"🟣 Direkt · Fehlt in Tour ({len(direkt_tour) if direkt_tour is not None else 0})")

    tabs = st.tabs(tab_labels)

    # --- Tab Übersicht ---
    with tabs[0]:
        ueb = _standort_uebersicht(hupa_sap, hupa_tour)
        st.markdown("**Pro Standort betroffene Kunden**")
        st.dataframe(ueb, use_container_width=True, hide_index=True)

        if hupa_sap is None or hupa_sap.empty:
            st.success("✅ Keine Hupa-Kunden mit Tagen in der Tourenplanung, die in SAP fehlen.")
        if direkt_sap is None or direkt_sap.empty:
            st.success("✅ Keine Direkt-Kunden mit Tagen in der Tourenplanung, die in SAP fehlen.")

    # --- Tab Hupa Fehlt in SAP ---
    with tabs[1]:
        if hupa_sap is None or hupa_sap.empty:
            st.info("Keine Treffer.")
        else:
            f1, f2 = st.columns([1, 2])
            standorte = ["Alle"] + sorted(hupa_sap["Standort"].unique().tolist())
            standort_wahl = f1.selectbox("Standort", standorte, key="filter_hupa_sap_standort")
            suche = f2.text_input("Suchen (SAP, Name, Straße, Ort)", key="filter_hupa_sap_suche")
            anz = _add_count_column(_filter_dataframe(hupa_sap, suche, standort_wahl))
            st.caption(f"{len(anz)} von {len(hupa_sap)} Zeilen")
            st.dataframe(anz, use_container_width=True, hide_index=True)

    # --- Tab Direkt Fehlt in SAP ---
    with tabs[2]:
        if direkt_sap is None or direkt_sap.empty:
            st.info("Keine Treffer.")
        else:
            suche = st.text_input("Suchen (SAP, Name, Straße, Ort)", key="filter_direkt_sap_suche")
            anz = _add_count_column(_filter_dataframe(direkt_sap, suche))
            st.caption(f"{len(anz)} von {len(direkt_sap)} Zeilen")
            st.dataframe(anz, use_container_width=True, hide_index=True)

    # --- Tabs Reverse (optional) ---
    if has_reverse:
        with tabs[3]:
            if hupa_tour is None or hupa_tour.empty:
                st.info("Keine Treffer.")
            else:
                f1, f2 = st.columns([1, 2])
                standorte = ["Alle"] + sorted(hupa_tour["Standort"].unique().tolist())
                standort_wahl = f1.selectbox("Standort", standorte, key="filter_hupa_tour_standort")
                suche = f2.text_input("Suchen (SAP, Name, Straße, Ort)", key="filter_hupa_tour_suche")
                anz = _add_count_column(_filter_dataframe(hupa_tour, suche, standort_wahl))
                st.caption(f"{len(anz)} von {len(hupa_tour)} Zeilen")
                st.dataframe(anz, use_container_width=True, hide_index=True)

        with tabs[4]:
            if direkt_tour is None or direkt_tour.empty:
                st.info("Keine Treffer.")
            else:
                suche = st.text_input("Suchen (SAP, Name, Straße, Ort)", key="filter_direkt_tour_suche")
                anz = _add_count_column(_filter_dataframe(direkt_tour, suche))
                st.caption(f"{len(anz)} von {len(direkt_tour)} Zeilen")
                st.dataframe(anz, use_container_width=True, hide_index=True)
